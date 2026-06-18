from io import BytesIO
import openpyxl

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from django.conf import settings
from django.db.models import Q
from django.core.paginator import Paginator
from django.core.mail import send_mail, EmailMessage

from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm

from rest_framework import viewsets

from .forms import RegisterForm
from .models import Product, Cart, ElemCart, Order, Manufacturer, Category

from .serializers import (
    ProductSerializer,
    CategorySerializer,
    ManufacturerSerializer,
    CartSerializer,
    ElemCartSerializer
)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .serializers import ProfileSerializer, OrderSerializer 
from .permissions import IsAdminOrReadOnly
from django.views.decorators.csrf import csrf_exempt

def index(request):
    popular_products = Product.objects.all().order_by('-id')[:6]
    categories = Category.objects.all()
    
    context = {
        'popular_products': popular_products,
        'categories': categories,
    }
    return render(request, 'main/index.html', context)

def product_list(request):
    query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    manufacturer_id = request.GET.get('manufacturer', '')
    
    products = Product.objects.all()
    
    if query:
        products = products.filter(Q(name__icontains=query) | Q(description__icontains=query))
    if category_id:
        products = products.filter(category_id=category_id)
    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()
    context = {
        'page_obj': page_obj,        
        'categories': categories,
        'manufacturers': manufacturers,
        'current_query': query,
        'current_category': category_id,
        'current_manufacturer': manufacturer_id,
    }
    return render(request, 'main/list.html',context,)

def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'main/product_detail.html', {'product': product})

@csrf_exempt
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_item, item_created = ElemCart.objects.get_or_create(cart=cart, product=product, defaults={'quantity': 1})
    
    if not item_created:
        cart_item.quantity += 1
        cart_item.save()
        
    return redirect('catalog:cart_view')

@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(ElemCart, id=item_id, cart__user=request.user)
    try:
        new_quantity = int(request.POST.get('quantity', 1))
    except (ValueError, TypeError):
        return HttpResponseBadRequest("Некорректное количество")
        
    if new_quantity > cart_item.product.stock:
        return HttpResponseBadRequest("Недостаточно товара на складе")
            
    cart_item.quantity = new_quantity
    cart_item.save()
    return redirect('catalog:cart_view')

@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(ElemCart, id=item_id, cart__user=request.user)
    cart_item.delete()
    return redirect('catalog:cart_view')

@login_required
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = ElemCart.objects.filter(cart=cart)
    total_cost = cart.total_price()
    return render(request, 'main/cart.html', {
        'cart_items': cart_items,
        'total_cost': total_cost
    })
@login_required
def checkout(request):
    cart = Cart.objects.get(user=request.user)
    cart_items = ElemCart.objects.filter(cart=cart)
    
    total = sum(item.quantity * item.product.price for item in cart_items)
    
    if not cart_items:
        return redirect('main:cart_view')
    
    if request.method == 'POST':
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        comment = request.POST.get('comment')

        order = Order.objects.create(
            user=request.user,
            total_price=total,
            status='pending'
        )

        for item in cart_items:
            ElemCart.objects.create(
                cart=cart,
                order=order,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price
            )
        send_receipt_email(request.user.email, order, cart_items, total)
        cart_items.delete()
        
        return redirect('main:order_success')
    
    return render(request, 'main/checkout.html', {
        'cart_items': cart_items,
        'total': total
    })

def generate_receipt(order, cart_items, total):
    """Простой Excel чек"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Чек_{order.id}"
    
    ws['A1'] = f'Чек №{order.id}'
    ws['A2'] = f'Дата: {order.created_at.strftime("%d.%m.%Y %H:%M")}'
    ws['A3'] = f'Покупатель: {order.user.username}'
    ws['A4'] = f'Email: {order.user.email}'
    ws['A5'] = f'Телефон: {order.phone}'
    ws['A6'] = f'Адрес: {order.address}'
    ws['A8'] = 'Товар'
    ws['B8'] = 'Кол-во'
    ws['C8'] = 'Цена'
    ws['D8'] = 'Сумма'
    
    row = 9
    for item in cart_items:
        ws[f'A{row}'] = item.product.name
        ws[f'B{row}'] = item.quantity
        ws[f'C{row}'] = float(item.product.price)
        ws[f'D{row}'] = float(item.quantity * item.product.price)
        row += 1
    
    ws[f'C{row}'] = 'ИТОГО:'
    ws[f'D{row}'] = float(total)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def send_receipt_email(recipient_email, order, excel_file):
    """Отправка чека на email"""
    subject = f'Чек по заказу №{order.id}'
    body = f'''
Здравствуйте, {order.user.username}!

Ваш заказ №{order.id} оформлен.
Сумма: {order.total_price} руб.
Телефон: {order.phone}
Адрес: {order.address}
Чек во вложении.
Спасибо за покупку!
'''
    email = EmailMessage(
        subject=subject,
        body=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[recipient_email]
    )
    email.attach(f'receipt_order_{order.id}.xlsx', excel_file.getvalue(),
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send()

def order_success(request):
    return render(request, 'main/order_success.html')

def send_receipt_email(recipient_email, order, cart_items, total):
    """Отправка чека на email с использованием send_mail"""
    
    subject = f'Чек по заказу №{order.id}'
    message = f"""
Здравствуйте, {order.user.username}!

Ваш заказ №{order.id} оформлен.
Дата: {order.created_at.strftime('%d.%m.%Y %H:%M')}
Сумма: {total} руб.
Телефон: {order.phone}
Адрес: {order.address}

Состав заказа:
"""
    for item in cart_items:
        message += f"- {item.product.name} x{item.quantity} = {item.quantity * item.product.price} руб.\n"
    
    message += f"""
Итого: {total} руб.

Спасибо за покупку!
"""
    send_mail(
        subject=subject,
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[recipient_email],
        fail_silently=False,
    )
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer

class ElemCartViewSet(viewsets.ModelViewSet):
    queryset = ElemCart.objects.all()
    serializer_class = ElemCartSerializer 

def register_view(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            login(request, user) 
            return redirect('catalog:index')
    else:
        form = RegisterForm()
    return render(request, 'main/register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('catalog:index')
    else:
        form = AuthenticationForm()
    return render(request, 'main/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('catalog:index')

@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def me_view(request):
    profile = request.user.profile
    if request.method == 'GET':
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)
        
    elif request.method == 'PATCH':
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly] 

class OrderViewSet(viewsets.ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if hasattr(user, 'profile') and user.profile.role == 'ADMIN':
            return Order.objects.all()
        return Order.objects.filter(user=user)
def profile_view(request):
    return render(request, 'main/profile.html')
def settings_view(request):
    return render(request, 'main/settings.html')
def test(request):
    return render(request, 'test.html')